import openpyxl
from datetime import datetime


def read_and_format_excel(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[4]]
    data = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        row_dict = dict(zip(headers, row))

        # Convert Resource Status → Active (Yes/No)
        if isinstance(row_dict.get("Resource Status"), bool):
            row_dict["Active"] = "Yes" if row_dict["Resource Status"] else "No"
        else:
            row_dict["Active"] = "Yes" if str(row_dict.get("Resource Status")).strip().lower() == "true" else "No"

        # Format date fields
        row_dict["Resource Start Date"] = format_excel_date(row_dict.get("Resource Start Date"))
        row_dict["Resource End Date"] = format_excel_date(row_dict.get("Resource End Date"))
        row_dict["Project Start Date"] = format_excel_date(row_dict.get("Project Start Date"))
        row_dict["Project End Date"] = format_excel_date(row_dict.get("Project End Date"))

        # Normalize Future Resource (Yes/No only, no "-")
        val = str(row_dict.get("Future Resource", "")).strip().lower()
        if val in ("yes", "y", "true", "1"):
            row_dict["Future Resource"] = "Yes"
        else:
            row_dict["Future Resource"] = "No"

        data.append({
            "Resource Id": str(int(row_dict.get("Resource Id"))) if row_dict.get("Resource Id") else "-",
            "Resource Name": row_dict.get("Resource Name"),
            "Active": row_dict.get("Active"),
            "Resource Start Date": row_dict.get("Resource Start Date"),
            "Resource End Date": row_dict.get("Resource End Date") or "-",
            "Project Name": row_dict.get("Project Name") or "-",
            "Project Start Date": row_dict.get("Project Start Date"),
            "Project End Date": row_dict.get("Project End Date") or "-",
            "Project Allocation": str(int(row_dict.get("Project Allocation"))) if row_dict.get("Project Allocation") else "-",
            "Future Resource": row_dict.get("Future Resource"),
        })
    return data


def format_excel_date(excel_date):
    """Convert Excel datetime or string to UI format dd-MMM-yyyy"""
    if not excel_date:
        return "-"
    if isinstance(excel_date, datetime):
        return excel_date.strftime("%d-%b-%Y")
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(excel_date), fmt).strftime("%d-%b-%Y")
        except ValueError:
            continue
    return str(excel_date)
